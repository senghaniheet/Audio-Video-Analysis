"""
Excel Loader Module
Handles Excel file creation and data loading
Excel format: mobile_number, order_id, customer_name, order_status, delivery_date, last_update
"""
import openpyxl
from openpyxl import Workbook
import os
from typing import Dict, Optional, List

# Excel file is now in mcp-order-status/server directory
EXCEL_FILE = "orders.xlsx"
EXCEL_FILE_LEGACY = "order_data.xlsx"  # Fallback for old format

def create_sample_excel():
    """Create Excel file with sample orders in new format"""
    from datetime import datetime, timedelta
    
    # Sample data with all required fields
    sample_orders = [
        {"mobile_number": "9876543210", "order_id": "MZ12345", "customer_name": "Rahul Sharma", "order_status": "Shipped", "delivery_date": "2025-11-20", "last_update": "2025-11-15"},
        {"mobile_number": "9988776655", "order_id": "AMZ22222", "customer_name": "Priya Patel", "order_status": "Out for Delivery", "delivery_date": "2025-11-18", "last_update": "2025-11-15"},
        {"mobile_number": "9123456780", "order_id": "AMZ33333", "customer_name": "Amit Verma", "order_status": "Delivered", "delivery_date": "2025-11-10", "last_update": "2025-11-10"},
        {"mobile_number": "9876543211", "order_id": "AMZ44444", "customer_name": "Sneha Kumar", "order_status": "Processing", "delivery_date": "2025-11-25", "last_update": "2025-11-14"},
        {"mobile_number": "9988776656", "order_id": "AMZ55555", "customer_name": "Vikram Singh", "order_status": "Shipped", "delivery_date": "2025-11-18", "last_update": "2025-11-15"},
        {"mobile_number": "9123456781", "order_id": "AMZ66666", "customer_name": "Anjali Desai", "order_status": "Out for Delivery", "delivery_date": "2025-11-17", "last_update": "2025-11-15"},
        {"mobile_number": "9876543212", "order_id": "AMZ77777", "customer_name": "Rajesh Mehta", "order_status": "Delivered", "delivery_date": "2025-11-12", "last_update": "2025-11-12"},
        {"mobile_number": "9988776657", "order_id": "AMZ88888", "customer_name": "Kavita Shah", "order_status": "Processing", "delivery_date": "2025-11-24", "last_update": "2025-11-13"},
        {"mobile_number": "9123456782", "order_id": "AMZ99999", "customer_name": "Mohit Agarwal", "order_status": "Shipped", "delivery_date": "2025-11-19", "last_update": "2025-11-15"},
        {"mobile_number": "9876543213", "order_id": "AMZ00001", "customer_name": "Divya Reddy", "order_status": "Out for Delivery", "delivery_date": "2025-11-16", "last_update": "2025-11-15"},
        {"mobile_number": "9988776658", "order_id": "AMZ00002", "customer_name": "Suresh Patel", "order_status": "Delivered", "delivery_date": "2025-11-11", "last_update": "2025-11-11"},
        {"mobile_number": "9123456783", "order_id": "AMZ00003", "customer_name": "Neha Verma", "order_status": "Processing", "delivery_date": "2025-11-23", "last_update": "2025-11-13"},
        {"mobile_number": "9876543214", "order_id": "AMZ00004", "customer_name": "Manoj Singh", "order_status": "Shipped", "delivery_date": "2025-11-21", "last_update": "2025-11-15"},
        {"mobile_number": "9988776659", "order_id": "AMZ00005", "customer_name": "Pooja Gupta", "order_status": "Out for Delivery", "delivery_date": "2025-11-17", "last_update": "2025-11-15"},
        {"mobile_number": "9123456784", "order_id": "MZ12345", "customer_name": "Rahul Sharma", "order_status": "Shipped", "delivery_date": "2025-11-20", "last_update": "2025-11-15"},
    ]
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Add headers in new format
    ws.append(["mobile_number", "order_id", "customer_name", "order_status", "delivery_date", "last_update"])
    
    # Add data rows
    for order in sample_orders:
        ws.append([
            order["mobile_number"],
            order["order_id"],
            order["customer_name"],
            order["order_status"],
            order["delivery_date"],
            order["last_update"]
        ])
    
    # Save file
    wb.save(EXCEL_FILE)
    print(f"Created {EXCEL_FILE} with {len(sample_orders)} sample orders")
    
    return EXCEL_FILE

def load_excel_data() -> List[Dict[str, str]]:
    """Load data from Excel file - supports new format and legacy formats"""
    
    excel_file = None
    
    # Try main Excel file first (orders.xlsx)
    if os.path.exists(EXCEL_FILE):
        excel_file = EXCEL_FILE
        print(f"[INFO] Using Excel file: {excel_file}")
    elif os.path.exists(EXCEL_FILE_LEGACY):
        excel_file = EXCEL_FILE_LEGACY
        print(f"[INFO] Using legacy Excel file: {excel_file}")
    else:
        print(f"[WARNING] No Excel file found. Creating sample file...")
        create_sample_excel()
        excel_file = EXCEL_FILE
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    orders = []
    
    # Get header row to determine format
    headers = [str(cell.value).lower() if cell.value else "" for cell in ws[1]]
    
    # Check for new format: mobile_number, order_id, customer_name, order_status, delivery_date, last_update
    is_new_format = "mobile_number" in headers and "customer_name" in headers
    # Check for backend format: MobileNumber, OrderID, CustomerName, etc.
    is_backend_format = "mobilenumber" in headers or "orderid" in headers
    
    # Skip header row (row 1)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:  # Skip if mobile/order_id is empty
            continue
            
        if is_new_format:
            # New format: mobile_number, order_id, customer_name, order_status, delivery_date, last_update
            orders.append({
                "mobile_number": str(row[0]).strip(),
                "order_id": str(row[1]).strip().upper(),
                "customer_name": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                "order_status": str(row[3]).strip() if len(row) > 3 and row[3] else "Unknown",
                "delivery_date": str(row[4]).strip() if len(row) > 4 and row[4] else "",
                "last_update": str(row[5]).strip() if len(row) > 5 and row[5] else ""
            })
        elif is_backend_format:
            # Backend format: MobileNumber, OrderID, CustomerName, OrderStatus, DeliveryDate, LastUpdate
            orders.append({
                "mobile_number": str(row[0]).strip(),
                "order_id": str(row[1]).strip().upper(),
                "customer_name": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                "order_status": str(row[3]).strip() if len(row) > 3 and row[3] else "Unknown",
                "delivery_date": str(row[4]).strip() if len(row) > 4 and row[4] else "",
                "last_update": str(row[5]).strip() if len(row) > 5 and row[5] else ""
            })
        else:
            # Legacy format: mobile_number, order_id, order_status (only 3 columns)
            orders.append({
                "mobile_number": str(row[0]).strip(),
                "order_id": str(row[1]).strip().upper(),
                "order_status": str(row[2]).strip() if len(row) > 2 and row[2] else "Unknown",
                "customer_name": "",
                "delivery_date": "",
                "last_update": ""
            })
    
    print(f"[OK] Loaded {len(orders)} orders from {excel_file}")
    return orders

def find_order(mobile_number: str, order_id: str) -> Optional[Dict[str, str]]:
    """Find order by mobile number and order ID"""
    
    orders = load_excel_data()
    
    # Clean inputs
    mobile_clean = str(mobile_number).strip().replace(" ", "").replace("-", "")
    order_id_clean = str(order_id).strip().upper().replace(" ", "")
    
    for order in orders:
        order_mobile = str(order["mobile_number"]).strip().replace(" ", "").replace("-", "")
        order_id_db = str(order["order_id"]).strip().upper().replace(" ", "")
        
        if order_mobile == mobile_clean and order_id_db == order_id_clean:
            return order
    
    return None

